#!/usr/bin/env python3
"""
NS 3420 PDF Import Tool
========================
Konverterer NS 3420 byggebeskrivelse-PDF til regneark (Excel/CSV)
og NS 3459 XML, kompatibelt med ns3420reader.

Bruk:
    python ns3420pdfimport.py beskrivelse.pdf
    python ns3420pdfimport.py beskrivelse.pdf --format all
    python ns3420pdfimport.py beskrivelse.pdf --format csv --output resultat
    python ns3420pdfimport.py extracted.txt --format xlsx

Kolonner (ns3420reader-format):
    Postnr ; NS3420 ; Emne ; Beskrivelse ; Mengde ; Enhet ; Pris ; Sum
"""

import re
import csv
import sys
import uuid
import subprocess
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple
import xml.etree.ElementTree as ET
from xml.dom import minidom

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XlSide
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ────────────────────────────────────────────────────────────
# Data models
# ────────────────────────────────────────────────────────────

@dataclass
class Post:
    """En NS 3420 post."""
    post_number: str = ""
    ns3420_code: str = ""
    subject: str = ""           # Emne (section title)
    description: str = ""       # Hovedbeskrivelse (CAPS title)
    quantity: float = 0.0       # Mengde
    unit: str = ""              # Enhet (m2, kg, stk, m, m3, RS)
    unit_price: Optional[float] = None   # Pris
    total_price: Optional[float] = None  # Sum
    full_text: str = ""         # Full spesifikasjonstekst
    chapter_code: str = ""
    chapter_title: str = ""
    section_code: str = ""
    section_title: str = ""
    location: str = ""          # Lokalisering
    page: int = 0


@dataclass
class Document:
    """Parsed NS 3420 dokument."""
    project_name: str = ""
    document_name: str = ""
    date: str = ""
    client: str = ""
    author: str = ""
    posts: List[Post] = field(default_factory=list)
    chapters: Dict[str, str] = field(default_factory=dict)


# ────────────────────────────────────────────────────────────
# Coordinate-based PDF parser (PyMuPDF)
# ────────────────────────────────────────────────────────────

try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False


class CoordinateParser:
    """Coordinate-based NS 3420 parser using PyMuPDF span positions.

    Column x-boundaries determined from PDF analysis (consistent across all 944 pages):
        Postnr:   [39, 99)
        Spec:     [99, 365)
        Enh:      [365, 397)
        Mengde:   [397, 460)
        Pris:     [460, 524)
        Sum:      [524, 580)
    """

    COL_RANGES = [
        ('postnr', 39, 99),
        ('spec', 99, 365),
        ('enh', 365, 397),
        ('mengde', 397, 460),
        ('pris', 460, 524),
        ('sum', 524, 580),
    ]
    HEADER_Y = 94
    FOOTER_Y = 762
    LINE_TOL = 4.0
    SUPER_SIZE = 9.0

    NS_CODE_RE = re.compile(r'^([A-Z]{2}\d+\.\d+[\w]*|[A-Z]{2,4}\d*[A-Z])$')
    CHAPTER_RE = re.compile(r'Kapittel:\s*(\d{2})\s+(.*)')
    POSTNR_LIKE = re.compile(r'^\d{2}\.\d')

    def parse(self, pdf_path: Path) -> Document:
        if not HAS_FITZ:
            raise ImportError("PyMuPDF (fitz) er ikke installert: pip install pymupdf")

        pdf = fitz.open(str(pdf_path))
        total = len(pdf)

        # Pass 1: extract raw rows from all pages
        raw_rows: List[dict] = []
        chapters: Dict[str, str] = {}
        project_name = ""
        date = ""
        author = ""
        client = ""

        for pi in range(total):
            if pi % 100 == 0:
                print(f"  Side {pi+1}/{total}...", end='\r')
            page = pdf[pi]
            hdr = self._parse_header(page)

            if hdr.get('chapter'):
                chapters[hdr['chapter'][0]] = hdr['chapter'][1]
            if hdr.get('project') and not project_name:
                project_name = hdr['project']
            if hdr.get('date') and not date:
                date = hdr['date']
            if hdr.get('author') and not author:
                author = hdr.get('author', '')
            if hdr.get('client') and not client:
                client = hdr.get('client', '')

            rows = self._extract_rows(page, pi + 1, hdr.get('chapter'))
            raw_rows.extend(rows)

        pdf.close()
        print(f"  {total} sider lest.              ")

        # Pass 2: merge split post numbers
        merged = self._merge_postnr(raw_rows)

        # Pass 3: build Post objects
        posts = self._build_posts(merged, chapters)

        return Document(
            project_name=project_name,
            document_name="NS 3420 Beskrivelse",
            date=date,
            author=author,
            client=client,
            posts=posts,
            chapters=chapters,
        )

    # ── Page header parsing ──

    def _parse_header(self, page) -> dict:
        result = {}
        blocks = page.get_text("dict").get("blocks", [])
        for block in blocks:
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    y = span["bbox"][1]
                    if y > self.HEADER_Y:
                        continue
                    txt = span["text"].strip()
                    m = self.CHAPTER_RE.search(txt)
                    if m:
                        result['chapter'] = (m.group(1), m.group(2).strip())
                    if 'Prosjekt:' in txt:
                        pm = re.search(r'Prosjekt:\s*(.+?)(?:Side|$)', txt)
                        if pm:
                            result['project'] = pm.group(1).strip()
                    dm = re.search(r'(\d{2}\.\d{2}\.\d{4})', txt)
                    if dm:
                        result['date'] = dm.group(1)
                    if re.match(r'^[A-ZÆØÅ][\w\s]+AS', txt):
                        result['author'] = txt
        return result

    # ── Row extraction ──

    def _col_for_x(self, x0: float) -> Optional[str]:
        for name, lo, hi in self.COL_RANGES:
            if lo <= x0 < hi:
                return name
        return None

    def _extract_rows(self, page, page_num: int, chapter_info=None) -> List[dict]:
        """Extract structured rows from a page using span coordinates."""
        text_dict = page.get_text("dict")

        # Collect content spans (skip header/footer)
        spans = []
        for block in text_dict.get("blocks", []):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    y0 = span["bbox"][1]
                    if y0 <= self.HEADER_Y or y0 >= self.FOOTER_Y:
                        continue
                    text = span["text"]
                    if not text.strip():
                        continue
                    spans.append({
                        'text': text,
                        'x0': span["bbox"][0],
                        'y0': y0,
                        'size': span["size"],
                        'font': span.get("font", ""),
                    })

        if not spans:
            return []

        # Group into lines by y-coordinate
        spans.sort(key=lambda s: (s['y0'], s['x0']))
        lines: List[List[dict]] = []
        cur_line = [spans[0]]
        cur_y = spans[0]['y0']

        for sp in spans[1:]:
            if abs(sp['y0'] - cur_y) <= self.LINE_TOL:
                cur_line.append(sp)
            else:
                lines.append(cur_line)
                cur_line = [sp]
                cur_y = sp['y0']
        lines.append(cur_line)

        # Convert to structured rows
        rows = []
        ch_code = chapter_info[0] if chapter_info else ""
        ch_title = chapter_info[1] if chapter_info else ""

        for line_spans in lines:
            row = self._line_to_row(line_spans, page_num, ch_code, ch_title)
            if row:
                rows.append(row)

        return rows

    KNOWN_UNITS = {'m', 'm2', 'm3', 'stk', 'kg', 'RS', 'tonn', '-', 'måned'}

    def _line_to_row(self, spans, page_num, ch_code, ch_title) -> Optional[dict]:
        """Assign spans to columns and build a row dict."""
        postnr = ""
        spec_parts = []
        enh_base = ""
        enh_super = ""
        mengde = ""
        pris = ""
        summ = ""

        for sp in sorted(spans, key=lambda s: s['x0']):
            x0 = sp['x0']
            text = sp['text'].strip()
            if not text:
                continue

            col = self._col_for_x(x0)
            if not col:
                continue

            # Fix: in the enh/mengde border zone [365,460), numeric values
            # are Mengde (right-aligned large numbers can start at x<397)
            if col == 'enh' and re.match(r'^[\d,.\s]+$', text) and len(text) > 2:
                col = 'mengde'
            # Also: if a span is in mengde range but looks like a unit, it's enh
            if col == 'mengde' and text in self.KNOWN_UNITS:
                col = 'enh'

            if col == 'postnr':
                # Only accept actual post numbers or digits in postnr column
                # Filter out description text that overflows into postnr column
                if re.match(r'^[\d.]+$', text):
                    postnr += text
                else:
                    # Text overflow from spec column - treat as spec
                    spec_parts.append(text)
            elif col == 'spec':
                spec_parts.append(text)
            elif col == 'enh':
                if sp['size'] < self.SUPER_SIZE:
                    enh_super += text  # superscript ², ³
                else:
                    enh_base += text
            elif col == 'mengde':
                mengde += text
            elif col == 'pris':
                pris += text
            elif col == 'sum':
                summ += text

        # Merge unit with superscript
        enh = enh_base + enh_super
        # Validate: if enh still has numbers glued to it, split them off
        if enh and not enh in self.KNOWN_UNITS:
            m = re.match(r'^([a-zA-Z]+\d?)([\d,.]+)$', enh)
            if m:
                unit_part = m.group(1)
                if unit_part in self.KNOWN_UNITS or unit_part + enh_super in self.KNOWN_UNITS:
                    mengde = m.group(2) + mengde
                    enh = unit_part + enh_super
                    enh_super = ""

        spec = ' '.join(spec_parts)
        postnr = postnr.strip()

        # Skip empty rows and known header text
        if not postnr and not spec and not enh and not mengde:
            return None
        skip_words = ('Postnr', 'Sum denne side', 'Akkumulert Kapittel')
        for sw in skip_words:
            if sw in (postnr + ' ' + spec):
                return None

        return {
            'postnr': postnr,
            'spec': spec,
            'enh': enh,
            'mengde': mengde,
            'pris': pris,
            'sum': summ,
            'page': page_num,
            'ch_code': ch_code,
            'ch_title': ch_title,
        }

    # ── Pass 2: merge split post numbers ──

    def _merge_postnr(self, rows: List[dict]) -> List[dict]:
        """Merge post numbers split across consecutive rows.

        Pattern A: postnr ends with '.' -> next postnr appended directly
            '09.235.10.' + '1.1' = '09.235.10.1.1'
        Pattern B: postnr is digits-only (1-3 chars) -> concatenated to previous
            '73.731.13' + '0' = '73.731.130'
        """
        if not rows:
            return rows

        merged: List[dict] = []
        i = 0
        while i < len(rows):
            row = dict(rows[i])  # copy

            # Check if next row is a continuation
            if i + 1 < len(rows) and row['postnr']:
                nxt = rows[i + 1]
                nxt_pn = nxt['postnr']

                if nxt_pn and self._is_continuation(row['postnr'], nxt_pn):
                    # Merge post number
                    if row['postnr'].endswith('.'):
                        row['postnr'] = row['postnr'] + nxt_pn
                    else:
                        row['postnr'] = row['postnr'] + nxt_pn

                    # Merge spec (append)
                    if nxt['spec']:
                        row['spec'] = (row['spec'] + '\n' + nxt['spec']).strip()

                    # Take enh/mengde from whichever row has them
                    if nxt['enh'] and not row['enh']:
                        row['enh'] = nxt['enh']
                    if nxt['mengde'] and not row['mengde']:
                        row['mengde'] = nxt['mengde']
                    if nxt['pris'] and not row['pris']:
                        row['pris'] = nxt['pris']
                    if nxt['sum'] and not row['sum']:
                        row['sum'] = nxt['sum']

                    i += 2  # skip the continuation row
                    merged.append(row)
                    continue

            merged.append(row)
            i += 1

        return merged

    def _is_continuation(self, prev_pn: str, next_pn: str) -> bool:
        """Is next_pn a continuation of prev_pn (not a standalone post)?"""
        # Pattern A: previous ends with dot -> next is appended directly
        if prev_pn.endswith('.'):
            return True
        # Pattern B: next is only digits (1-3 chars) -> concatenation (e.g., 73.731.13 + 0 = 73.731.130)
        if re.match(r'^\d{1,3}$', next_pn) and self.POSTNR_LIKE.match(prev_pn):
            return True
        # Pattern C: next is N.M format (e.g., 8.1, 9.2, 2.1) -> sub-post continuation
        # Used by ch73: 73.731.19 + 8.1 = 73.731.198.1
        if re.match(r'^\d+\.\d+$', next_pn) and self.POSTNR_LIKE.match(prev_pn):
            return True
        return False

    # ── Pass 3: build posts ──

    def _build_posts(self, rows: List[dict], chapters: Dict[str, str]) -> List[Post]:
        posts: List[Post] = []
        current_post: Optional[Post] = None
        current_section = ("", "")

        for row in rows:
            postnr = row['postnr']
            spec = row['spec']
            enh = row['enh']
            mengde_str = row['mengde']

            # Check for section heading (2-level like "05.21" + title)
            if postnr and re.match(r'^\d{2}\.\d{1,3}$', postnr) and spec and not enh:
                # Could be section heading OR a post (kap 25/26/72 have 2-level posts)
                # It's a post if the spec starts with an NS code
                first_word = spec.split()[0] if spec.split() else ""
                if self.NS_CODE_RE.match(first_word):
                    pass  # It's a 2-level post, fall through to post creation
                else:
                    current_section = (postnr, spec.split('\n')[0][:60])
                    # Also emit as a context row
                    posts.append(Post(
                        description=f"{postnr} {spec}",
                        chapter_code=row['ch_code'],
                        chapter_title=row['ch_title'],
                        subject=spec.split('\n')[0][:60],
                        page=row['page'],
                    ))
                    continue

            # Chapter heading (single 2-digit code)
            if postnr and re.match(r'^\d{2}$', postnr) and spec and not enh:
                current_section = ("", "")
                posts.append(Post(
                    description=f"{postnr} {spec}",
                    chapter_code=row['ch_code'],
                    chapter_title=row['ch_title'],
                    subject=spec.split('\n')[0][:60],
                    page=row['page'],
                ))
                continue

            # Context text (no postnr, just descriptive text between posts)
            if not postnr and spec and not enh and not mengde_str:
                if current_post:
                    current_post.description += '\n' + spec
                else:
                    posts.append(Post(
                        description=spec,
                        chapter_code=row['ch_code'],
                        chapter_title=row['ch_title'],
                        subject=current_section[1] if current_section[1] else row['ch_title'],
                        page=row['page'],
                    ))
                continue

            # Row with data but no postnr -> continuation of current post
            if not postnr:
                if current_post:
                    if spec:
                        current_post.description += '\n' + spec
                    if enh and not current_post.unit:
                        current_post.unit = enh
                    if mengde_str and current_post.quantity == 0:
                        current_post.quantity = self._parse_qty(mengde_str)
                continue

            # Row with postnr -> new post or sub-item
            if self.POSTNR_LIKE.match(postnr):
                # Save previous post
                if current_post:
                    posts.append(current_post)

                # Extract NS code from spec
                ns_code, desc = self._split_ns_code(spec)

                current_post = Post(
                    post_number=postnr,
                    ns3420_code=ns_code,
                    description=desc,
                    unit=enh,
                    quantity=self._parse_qty(mengde_str),
                    chapter_code=row['ch_code'],
                    chapter_title=row['ch_title'],
                    section_code=current_section[0],
                    section_title=current_section[1],
                    subject=current_section[1] if current_section[1] else row['ch_title'],
                    page=row['page'],
                )
            else:
                # Postnr doesn't look like a standard number
                # Treat as description continuation
                if current_post:
                    current_post.description += '\n' + postnr + ' ' + spec
                    if enh and not current_post.unit:
                        current_post.unit = enh
                    if mengde_str and current_post.quantity == 0:
                        current_post.quantity = self._parse_qty(mengde_str)

        # Don't forget the last post
        if current_post:
            posts.append(current_post)

        # Post-process: fix RS quantities, clean descriptions
        for p in posts:
            if p.unit == 'RS' and p.quantity == 0:
                p.quantity = 1.0
            # Clean up description whitespace
            p.description = re.sub(r'\n{3,}', '\n\n', p.description).strip()

        return posts

    def _split_ns_code(self, spec: str) -> Tuple[str, str]:
        """Extract NS code from beginning of spec text."""
        if not spec:
            return ("", "")
        first_line = spec.split('\n')[0]
        words = first_line.split()
        if words and self.NS_CODE_RE.match(words[0]):
            code = words[0]
            desc = first_line[len(code):].strip()
            if '\n' in spec:
                desc += '\n' + '\n'.join(spec.split('\n')[1:])
            return (code, desc)
        return ("", spec)

    @staticmethod
    def _parse_qty(s: str) -> float:
        if not s:
            return 0.0
        try:
            return float(s.strip().replace(' ', '').replace(',', '.'))
        except ValueError:
            return 0.0


# ────────────────────────────────────────────────────────────
# Text extraction (fallback for .txt files)
# ────────────────────────────────────────────────────────────

EXTRACTOR_DATA_DIR = Path.home() / "Dokumenter/Prosjekt/Rust/extractor-data"


def find_existing_extraction(pdf_path: Path) -> Optional[Path]:
    """Ser etter eksisterende extractor-data output."""
    output_dir = EXTRACTOR_DATA_DIR / "output"
    candidates = [
        output_dir / f"{pdf_path.stem}.txt",
        pdf_path.with_suffix('.txt'),
        pdf_path.parent / f"{pdf_path.stem}.txt",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def extract_with_extractor_data(pdf_path: Path) -> Optional[str]:
    """Kjor extractor-data ai_convert binary."""
    binary = EXTRACTOR_DATA_DIR / "target/release/ai_convert"
    if not binary.exists():
        return None

    print(f"  Kjorer extractor-data ai_convert...")
    try:
        result = subprocess.run(
            [str(binary), str(pdf_path)],
            capture_output=True, text=True, timeout=600,
            cwd=str(EXTRACTOR_DATA_DIR)
        )
        if result.returncode == 0:
            output_file = EXTRACTOR_DATA_DIR / "output" / f"{pdf_path.stem}.txt"
            if output_file.exists():
                return output_file.read_text(encoding='utf-8')
            return result.stdout
        else:
            print(f"  extractor-data feilet: {result.stderr[:200]}")
    except subprocess.TimeoutExpired:
        print("  extractor-data tidsavbrudd (>600s)")
    except FileNotFoundError:
        pass
    return None


def extract_with_pymupdf(pdf_path: Path) -> Optional[str]:
    """Ekstraher tekst med PyMuPDF (fitz)."""
    try:
        import fitz
    except ImportError:
        return None

    print("  Bruker PyMuPDF for tekstekstraksjon...")
    doc = fitz.open(str(pdf_path))
    pages = []
    for i in range(len(doc)):
        page = doc[i]
        text = page.get_text("text")
        pages.append(f"\n=== Side {i+1} ===\n{text}")
    doc.close()
    return "\n".join(pages)


def extract_with_pdfplumber(pdf_path: Path) -> Optional[str]:
    """Ekstraher tekst med pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        return None

    print("  Bruker pdfplumber for tekstekstraksjon...")
    pages = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            pages.append(f"\n=== Side {i+1} ===\n{text}")
    return "\n".join(pages)


def extract_text(pdf_path: Path) -> str:
    """Ekstraher tekst fra PDF med beste tilgjengelige metode."""
    existing = find_existing_extraction(pdf_path)
    if existing:
        print(f"  Bruker eksisterende ekstrahert fil: {existing}")
        return existing.read_text(encoding='utf-8')

    for method in [extract_with_extractor_data, extract_with_pymupdf, extract_with_pdfplumber]:
        result = method(pdf_path)
        if result:
            return result

    print("FEIL: Ingen PDF-leser tilgjengelig.")
    print("Installer en av:")
    print("  pip install pymupdf")
    print("  pip install pdfplumber")
    print("Eller bygg extractor-data:")
    print(f"  cd {EXTRACTOR_DATA_DIR} && cargo build --release --bin ai_convert")
    sys.exit(1)


# ────────────────────────────────────────────────────────────
# Parser
# ────────────────────────────────────────────────────────────

class NS3420Parser:
    """Parser for ekstrahert NS 3420 tekst."""

    # Page marker
    PAGE_RE = re.compile(r'=== Side (\d+) ===')
    # Chapter from page header: "Kapittel: 05 Betongarbeider"
    CHAPTER_RE = re.compile(r'Kapittel:\s*(\d{2})\s+(.*)')

    # Post number: minimum 3 levels (XX.XX.N or XX.XXX.N)
    PN = r'(\d{2}\.\d{2,3}\.\d+(?:\.\d+)*)'
    # NS3420 code: 2 uppercase letters + digits + dot + digits + optional suffix
    # Also handles short codes like WZA, WRA, WR2A (without dot)
    NC = r'([A-Z]{2}\d+\.\d+[\w]*|[A-Z]{2,4}\d*[A-Z])'

    # Table format: | post_num | ns_code ... |
    TABLE_POST_RE = re.compile(rf'\|\s*{PN}\s*\|\s*{NC}')
    # Compact table: | post_num ns_code ... |
    COMPACT_POST_RE = re.compile(rf'\|\s*{PN}\s+{NC}')
    # Plain text: post_num ns_code (space-separated)
    PLAIN_POST_RE = re.compile(rf'^{PN}\s+{NC}')

    # Dot-connected: POSTNR.NSCODE (e.g., 09.235.10.NB2.162A, 15.244.11.SB1.8221)
    DOT_POST_RE = re.compile(rf'{PN}\.{NC}')
    # Same in table cell
    TABLE_DOT_RE = re.compile(rf'\|\s*{PN}\.{NC}')

    # 2-level post numbers (e.g., 25.21, 72.16) - used by kap 25, 26, 72 etc.
    PN2 = r'(\d{2}\.\d{1,3})'
    TABLE_2LVL_RE = re.compile(rf'\|\s*{PN2}\s*\|\s*{NC}')
    COMPACT_2LVL_RE = re.compile(rf'\|\s*{PN2}\s+{NC}')
    PLAIN_2LVL_RE = re.compile(rf'^{PN2}\s+{NC}')
    DOT_2LVL_RE = re.compile(rf'{PN2}\.{NC}')

    # Quantity/unit patterns (applied to cleaned text, order matters - first match wins)
    QTY_PATTERNS = [
        # Antallstk N
        (re.compile(r'Antall\s*stk\s+([\d,.]+)', re.I), 'stk'),
        # Areal (various forms) ...m N (m2)
        (re.compile(r'[Aa]real[^|\n]*?m\s+([\d,.]+)'), 'm2'),
        # Lengde / Samlet lengde / Lengde på tettet fuge ...m N
        (re.compile(r'(?:[Ss]amlet\s+)?[Ll]engde[^|\n]*?m\s+([\d,.]+)'), 'm'),
        # Massekg N
        (re.compile(r'[Mm]asse\s*kg\s+([\d,.]+)'), 'kg'),
        # Volum m N (m3)
        (re.compile(r'[Vv]olum\s*m\s+([\d,.]+)'), 'm3'),
        # Rund sum RS
        (re.compile(r'Rund\s+sum\s*RS', re.I), 'RS'),
        # m2m N (m2 written as text)
        (re.compile(r'm2m\s+([\d,.]+)'), 'm2'),
        # Breddem N (width in meters)
        (re.compile(r'[Bb]redde\s*m\s+([\d,.]+)'), 'm'),
        # Prosjektert fast volum m3
        (re.compile(r'[Pp]rosjektert fast volum\s*m\s+([\d,.]+)'), 'm3'),
        # "stk N" standalone in table
        (re.compile(r'\bstk\s+([\d,.]+)', re.I), 'stk'),
    ]

    # Lokalisering
    LOK_RE = re.compile(r'Lokalisering:\s*(.+?)(?:\||$|\n)')

    def __init__(self):
        self.reset()

    def reset(self):
        self.posts: List[Post] = []
        self.chapters: Dict[str, str] = {}
        self._ch = ("", "")       # current chapter (code, title)
        self._sec = ("", "")      # current section (code, title)
        self._page = 0
        self.project_name = ""
        self.date = ""
        self.client = ""
        self.author = ""

    @staticmethod
    def _preprocess(text: str) -> str:
        """Fjern kolonnemarkeringer og normaliser tekst."""
        lines = []
        for line in text.split('\n'):
            stripped = line.strip()
            # Skip column layout markers
            if stripped in ('[Venstre kolonne]', '[Høyre kolonne]'):
                continue
            # Strip leading whitespace from indented content (column layout)
            lines.append(stripped if stripped else line)
        return '\n'.join(lines)

    def parse(self, text: str) -> Document:
        self.reset()
        text = self._preprocess(text)
        lines = text.split('\n')
        self._extract_metadata(lines)
        self._parse_lines(lines)
        return Document(
            project_name=self.project_name,
            document_name="NS 3420 Beskrivelse",
            date=self.date,
            client=self.client,
            author=self.author,
            posts=self.posts,
            chapters=self.chapters,
        )

    # ── Metadata ──

    def _extract_metadata(self, lines: List[str]):
        for line in lines[:100]:
            if 'Prosjekt:' in line:
                m = re.search(r'Prosjekt:\s*(.+?)(?:Side|$)', line)
                if m and not self.project_name:
                    self.project_name = m.group(1).strip()
            m = re.search(r'(\d{2}\.\d{2}\.\d{4})', line)
            if m and not self.date:
                self.date = m.group(1)
            if line.startswith('## ') or line.startswith('# '):
                clean = line.lstrip('#').strip()
                if re.match(r'^[A-ZÆØÅ][\w\s]+AS', clean) and not self.author:
                    self.author = clean
                elif 'Eiendom' in clean and not self.client:
                    self.client = clean

    # ── Context tracking ──

    def _is_skip_line(self, line: str) -> bool:
        s = line.strip()
        if not s:
            return False
        if re.match(r'^[A-ZÆØÅ][\w\s]+AS', s) or s.startswith('Prosjekt:'):
            return True
        if 'Postnr:::NS-kode' in s:
            return True
        if s.startswith('Sum denne side') or s.startswith('Akkumulert Kapittel'):
            return True
        if s == 'INNHOLDSFORTEGNELSE':
            return True
        return False

    def _update_context(self, line: str):
        m = self.PAGE_RE.search(line)
        if m:
            self._page = int(m.group(1))
            return

        m = self.CHAPTER_RE.search(line)
        if m:
            code, title = m.group(1), m.group(2).strip()
            self._ch = (code, title)
            self.chapters[code] = title
            return

        # Section headings: XX.XX or XX.XXX (2-level only)
        stripped = line.strip().lstrip('|').strip()
        m = re.match(r'^(\d{2}\.\d{2,3})\s*\|?\s+([A-Za-zÆØÅæøå][\w\sæøåÆØÅ,.()-]+)', stripped)
        if m:
            code = m.group(1)
            if len(code.split('.')) == 2:
                self._sec = (code, m.group(2).strip().rstrip('|').strip())

    # ── Post detection ──

    def _detect_post(self, line: str) -> Optional[Tuple[str, str]]:
        """Returnerer (postnr, ns_kode) eller None."""
        # 1. Table: | post_num | ns_code |
        m = self.TABLE_POST_RE.search(line)
        if m:
            return (m.group(1), m.group(2))
        # 2. Compact table: | post_num ns_code |
        m = self.COMPACT_POST_RE.search(line)
        if m:
            return (m.group(1), m.group(2))
        # 3. Plain text: post_num ns_code
        m = self.PLAIN_POST_RE.match(line.strip())
        if m:
            return (m.group(1), m.group(2))
        # 4. Dot-connected in table: | POSTNR.NSCODE |
        m = self.TABLE_DOT_RE.search(line)
        if m:
            return (m.group(1), m.group(2))
        # 5. Dot-connected plain: POSTNR.NSCODE
        stripped = line.strip()
        m = self.DOT_POST_RE.match(stripped)
        if m:
            return (m.group(1), m.group(2))
        # 6. 2-level post numbers (kap 25, 26, 72 etc.): | XX.N | NSCODE |
        m = self.TABLE_2LVL_RE.search(line)
        if m:
            return (m.group(1), m.group(2))
        # 7. 2-level compact: | XX.N NSCODE |
        m = self.COMPACT_2LVL_RE.search(line)
        if m:
            return (m.group(1), m.group(2))
        # 8. 2-level plain: XX.N NSCODE
        m = self.PLAIN_2LVL_RE.match(stripped)
        if m:
            return (m.group(1), m.group(2))
        # 9. 2-level dot-connected: XX.N.NSCODE
        m = self.DOT_2LVL_RE.match(stripped)
        if m:
            return (m.group(1), m.group(2))
        # 10. Post number only (no NS code) - plain text
        # e.g., "12.251.1 Her medtas arbeider med..."
        m = re.match(r'^(\d{2}\.\d{2,3}\.\d+(?:\.\d+)*)\s+([A-ZÆØÅa-zæøå])', stripped)
        if m:
            word2 = stripped.split()[1] if len(stripped.split()) > 1 else ''
            if not re.match(r'^[A-Z]{2}\d', word2):
                return (m.group(1), '')
        # 11. Post number in table cell, no NS code
        # e.g., "| 05.24.8.1 B x L x T = ... |"
        m = re.match(r'\|\s*(\d{2}\.\d{2,3}\.\d+(?:\.\d+)*)\s+([A-ZÆØÅa-zæøå])', line)
        if m:
            word2 = line.split()[1] if len(line.split()) > 1 else ''
            # Make sure it's not an NS code
            after_pn = line[m.end(1):].strip().lstrip('|').strip()
            first_word = after_pn.split()[0] if after_pn.split() else ''
            if not re.match(r'^[A-Z]{2}\d', first_word):
                return (m.group(1), '')
        return None

    # ── Main parse loop ──

    def _is_context_line(self, line: str) -> bool:
        """Check if line is meaningful non-post text (section header, context)."""
        stripped = line.strip().replace('|', ' ').strip()
        stripped = re.sub(r'-{3,}', '', stripped).strip()
        if not stripped or len(stripped) < 3:
            return False
        # Chapter/section headings like "05 Betongarbeider" or "05.21 Grunn og fundamenter"
        if re.match(r'^\d{2}(?:\.\d{1,3})?\s+[A-Za-zÆØÅæøå]', stripped):
            return True
        # Descriptive text like "Verkstedhall - Smøregraver"
        if re.match(r'^[A-ZÆØÅa-zæøå][\w\sæøåÆØÅ,.()-]+$', stripped) and len(stripped) > 5:
            return True
        return False

    def _parse_lines(self, lines: List[str]):
        current_post: Optional[Post] = None
        post_lines: List[str] = []
        found_first_chapter = False
        pending_context: List[str] = []  # Non-post text between posts

        for line in lines:
            raw = line.rstrip()

            # Wait for first chapter header (skip TOC)
            if not found_first_chapter:
                if self.CHAPTER_RE.search(raw):
                    found_first_chapter = True
                    self._update_context(raw)
                continue

            self._update_context(raw)

            if self._is_skip_line(raw):
                continue

            post_info = self._detect_post(raw)
            if post_info:
                # Save previous post
                if current_post:
                    self._finalize_post(current_post, post_lines)
                    self.posts.append(current_post)

                # Flush pending context as a text-only row
                if pending_context:
                    ctx_text = '\n'.join(
                        re.sub(r'\s{2,}', ' ', re.sub(r'-{3,}', '', re.sub(r'\|', ' ', l))).strip()
                        for l in pending_context if l.strip()
                    )
                    if ctx_text.strip():
                        self.posts.append(Post(
                            description=ctx_text.strip(),
                            chapter_code=self._ch[0],
                            chapter_title=self._ch[1],
                            section_code=self._sec[0],
                            section_title=self._sec[1],
                            subject=self._sec[1] or self._ch[1],
                            page=self._page,
                        ))
                    pending_context = []

                pn, code = post_info
                current_post = Post(
                    post_number=pn,
                    ns3420_code=code,
                    chapter_code=self._ch[0],
                    chapter_title=self._ch[1],
                    section_code=self._sec[0],
                    section_title=self._sec[1],
                    page=self._page,
                )
                post_lines = [raw]
            elif current_post:
                # Don't add page/chapter headers to post text
                if not self.PAGE_RE.search(raw) and not self.CHAPTER_RE.search(raw):
                    post_lines.append(raw)
            else:
                # No current post - capture as context if meaningful
                if self._is_context_line(raw):
                    pending_context.append(raw)

        if current_post:
            self._finalize_post(current_post, post_lines)
            self.posts.append(current_post)

        # Second pass: extract sub-posts from within parent post full_text
        self._extract_sub_posts()

    # ── Post finalization ──

    def _clean_lines(self, lines: List[str]) -> List[str]:
        """Remove table formatting from lines."""
        clean = []
        for line in lines:
            cl = re.sub(r'\|', ' ', line)
            cl = re.sub(r'-{3,}', '', cl)
            cl = cl.strip()
            if cl:
                clean.append(cl)
        return clean

    def _build_full_description(self, post: Post, clean_lines: List[str]) -> str:
        """Build full specification text for Beskrivelse field.

        Includes NS-code, CAPS title, all key-value fields,
        Andre krav, etc. - everything in the NS-kode/Spesifikasjon column.
        """
        desc_parts = []
        skip_first = True
        for cl in clean_lines:
            # Skip the first line (post number + NS code) - that info is in separate columns
            if skip_first:
                if post.post_number in cl or post.ns3420_code in cl:
                    # Check if there's description text after the code on same line
                    idx = cl.find(post.ns3420_code)
                    if idx >= 0:
                        after = cl[idx + len(post.ns3420_code):].strip()
                        if after:
                            desc_parts.append(after)
                    skip_first = False
                    continue
                # If post_number wasn't on first line, include all
                skip_first = False

            # Skip standalone superscript numbers (2, 3 from m², m³)
            if cl.strip() in ('2', '3') and len(cl.strip()) == 1:
                continue

            # Normalize whitespace
            cl = re.sub(r'\s{2,}', ' ', cl)
            desc_parts.append(cl)

        return '\n'.join(desc_parts)

    def _finalize_post(self, post: Post, lines: List[str]):
        raw_text = '\n'.join(lines)
        post.full_text = raw_text

        clean_lines = self._clean_lines(lines)
        clean_text = '\n'.join(clean_lines)

        # Detect sub-item number from next line (e.g., "1.1 GLIDESJIKT")
        self._detect_sub_number(post, clean_lines)

        # Build full description (all spec text)
        post.description = self._build_full_description(post, clean_lines)

        self._extract_quantity(post, clean_text)
        self._extract_location(post, clean_text)

        post.subject = post.section_title or post.chapter_title

    # Sub-post pattern: "POSTNR. Description\nN    Antallstk XX"
    SUBPOST_RE = re.compile(
        r'(?:^|\|)\s*(\d{2}\.\d{2,3}\.\d+(?:\.\d+)*)\.\s*'  # parent postnr with trailing dot
        r'(.+?)$',  # description text
        re.MULTILINE
    )
    # Sub-number + quantity in next lines
    SUBQTY_PATTERNS = [
        re.compile(r'Antall\s*stk\s*([\d,.]+)', re.I),
        re.compile(r'Areal[^|\n]*?m\s*([\d,.]+)', re.I),
        re.compile(r'(?:Samlet\s+)?[Ll]engde[^|\n]*?m\s*([\d,.]+)'),
        re.compile(r'Masse\s*kg\s*([\d,.]+)', re.I),
        re.compile(r'Volum\s*m\s*([\d,.]+)', re.I),
    ]
    SUBUNIT_MAP = {0: 'stk', 1: 'm2', 2: 'm', 3: 'kg', 4: 'm3'}

    def _extract_sub_posts(self):
        """Extract sub-posts embedded within parent post text.

        Patterns found in NS 3420:
        1. Table: | 05.24.20.1 Inntil Ø100 | ... | Antallstk 2 |
        2. Table: | 06.24.1.1 | L x H x T = ... | Arealm 2 50,00 |
        3. Plain: 03.034.13.\n1\n  Antallstk 26
        4. Compact: | POSTNR. Description | qty |
        """
        new_posts = []

        for parent in self.posts:
            if not parent.post_number:
                continue

            text = parent.full_text
            parent_pn = re.escape(parent.post_number)

            # Clean the full text for searching
            clean = re.sub(r'\|', ' ', text)
            clean = re.sub(r'-{3,}', ' ', clean)

            # Pattern A: "PARENT.N description" on ONE line
            # Matches: 05.24.20.1 Inntil Ø100, 06.24.1.1 L x H x T, etc.
            sub_re = re.compile(
                rf'{parent_pn}\.(\d+(?:\.\d+)*)\s*(.*)',
                re.MULTILINE
            )

            # Pattern B: "PARENT. description" then "N" on NEXT line (split format)
            # Matches: 03.034.13. For trekkekum\n1\nAntallstk 26
            split_re = re.compile(
                rf'{parent_pn}\.\s*([^\d\n].*?)$',
                re.MULTILINE
            )

            # Pass 1: Direct matches (PARENT.N on same line)
            for m in sub_re.finditer(clean):
                sub_num = m.group(1)
                full_pn = f"{parent.post_number}.{sub_num}"
                desc_text = m.group(2).strip()

                context_start = m.start()
                context_end = min(len(clean), m.end() + 300)
                context = clean[context_start:context_end]

                # Extract quantity/unit from context
                qty = 0.0
                unit = ""
                for pi, pat in enumerate(self.SUBQTY_PATTERNS):
                    qm = pat.search(context)
                    if qm:
                        try:
                            qty = float(qm.group(1).replace(',', '.').replace(' ', ''))
                            unit = self.SUBUNIT_MAP[pi]
                            break
                        except ValueError:
                            continue

                # Also check for "Arealm 2 N" pattern (m² with split superscript)
                if not unit:
                    am = re.search(r'Areal\s*m\s*2\s+([\d,.]+)', context)
                    if am:
                        try:
                            qty = float(am.group(1).replace(',', '.'))
                            unit = 'm2'
                        except ValueError:
                            pass

                # Clean up description
                desc_text = re.sub(r'\s{2,}', ' ', desc_text)[:200]

                new_posts.append(Post(
                    post_number=full_pn,
                    ns3420_code=parent.ns3420_code,
                    subject=parent.subject,
                    description=desc_text if desc_text else parent.description,
                    quantity=qty,
                    unit=unit,
                    chapter_code=parent.chapter_code,
                    chapter_title=parent.chapter_title,
                    section_code=parent.section_code,
                    section_title=parent.section_title,
                    page=parent.page,
                ))

            # Pass 2: Split matches (PARENT. text\nN\n qty)
            for m in split_re.finditer(clean):
                desc_text = m.group(1).strip()
                # Look for sub-number on subsequent lines
                after = clean[m.end():m.end() + 200]
                lines_after = after.strip().split('\n')
                for la in lines_after[:3]:
                    la = la.strip()
                    num_m = re.match(r'^(\d{1,3})\s', la)
                    if num_m:
                        sub_num = num_m.group(1)
                        full_pn = f"{parent.post_number}.{sub_num}"
                        # Get context for quantity extraction
                        context = '\n'.join(lines_after[:5])

                        qty = 0.0
                        unit = ""
                        for pi, pat in enumerate(self.SUBQTY_PATTERNS):
                            qm = pat.search(context)
                            if qm:
                                try:
                                    qty = float(qm.group(1).replace(',', '.').replace(' ', ''))
                                    unit = self.SUBUNIT_MAP[pi]
                                    break
                                except ValueError:
                                    continue
                        if not unit:
                            am = re.search(r'Areal\s*m\s*2\s+([\d,.]+)', context)
                            if am:
                                try:
                                    qty = float(am.group(1).replace(',', '.'))
                                    unit = 'm2'
                                except ValueError:
                                    pass

                        new_posts.append(Post(
                            post_number=full_pn,
                            ns3420_code=parent.ns3420_code,
                            subject=parent.subject,
                            description=desc_text[:200] if desc_text else parent.description,
                            quantity=qty,
                            unit=unit,
                            chapter_code=parent.chapter_code,
                            chapter_title=parent.chapter_title,
                            section_code=parent.section_code,
                            section_title=parent.section_title,
                            page=parent.page,
                        ))
                        break  # Only first sub-number per split match

        # Deduplicate: prefer posts with quantity over those without
        existing = {p.post_number for p in self.posts}
        for np in new_posts:
            if np.post_number not in existing:
                self.posts.append(np)
                existing.add(np.post_number)

        # Sort by post number
        self.posts.sort(key=lambda p: self._sort_key(p.post_number))

    @staticmethod
    def _sort_key(postnr: str) -> List:
        """Natural sort key for post numbers."""
        parts = postnr.split('.')
        result = []
        for p in parts:
            try:
                result.append((0, int(p)))
            except ValueError:
                result.append((1, p))
        return result

    def _detect_sub_number(self, post: Post, clean_lines: List[str]):
        """Detect sub-item number and append to post number.

        Handles:
        - '1.1 GLIDESJIKT'  -> sub-post, append '.1.1'
        - '0 GEOTEKSTIL...' -> continuation digit, append '0' (no dot)
        - standalone '2'/'3' followed by 'Areal'/'Volum' -> m²/m³ superscript, SKIP
        """
        for i, cl in enumerate(clean_lines[1:4], 1):
            stripped = cl.strip()
            # Pattern 1: "N.M DESCRIPTION" -> sub-post number (with dots)
            m = re.match(r'^(\d+\.\d+)\s+[A-ZÆØÅ]', stripped)
            if m:
                sub = m.group(1)
                post.post_number = f"{post.post_number}.{sub}"
                return

            # Standalone digit: check context to decide what it means
            if re.match(r'^\d{1,2}$', stripped):
                digit = stripped
                # Check next line
                next_line = clean_lines[i + 1].strip() if i + 1 < len(clean_lines) else ""
                next_lower = next_line.lower()
                # Skip if it's a m²/m³ superscript (digit 2/3 before Areal/Volum/Prosjektert)
                if digit in ('2', '3') and any(
                    next_lower.startswith(w)
                    for w in ('areal', 'volum', 'prosjektert', 'forskalings', 'm2', 'm3')
                ):
                    continue
                # Otherwise it's a continuation of the post number
                alpha = re.sub(r'[^a-zA-ZæøåÆØÅ]', '', next_line)
                is_caps = len(alpha) >= 3 and sum(1 for c in alpha if c.isupper()) / len(alpha) > 0.6
                if is_caps:
                    post.post_number = f"{post.post_number}{digit}"
                    return

            # Pattern 3: "N CAPS_DESCRIPTION" where N is single digit -> continuation
            # E.g., 73.731.13 + "0 GEOTEKSTIL" = 73.731.130
            m = re.match(r'^(\d)\s+[A-ZÆØÅ]{3,}', stripped)
            if m:
                digit = m.group(1)
                # Get the text after the digit
                desc_after = stripped[len(digit):].strip()
                # Skip if this looks like a quantity line (not a postnr continuation)
                if re.match(r'^(Antall|Areal|Lengde|Masse|Volum|Rund)', desc_after, re.I):
                    continue
                post.post_number = f"{post.post_number}{digit}"
                return

    def _extract_quantity(self, post: Post, clean_text: str):
        """Ekstraher mengde og enhet."""
        for pattern, unit in self.QTY_PATTERNS:
            m = pattern.search(clean_text)
            if m:
                if unit == 'RS':
                    post.unit = 'RS'
                    post.quantity = 1.0
                    return
                try:
                    qty_str = m.group(1).strip().replace(' ', '').replace(',', '.')
                    val = float(qty_str)
                    if val > 0:
                        post.quantity = val
                        post.unit = unit
                        return
                except (ValueError, IndexError):
                    continue

    def _extract_location(self, post: Post, clean_text: str):
        m = self.LOK_RE.search(clean_text)
        if m:
            loc = m.group(1).strip()
            # Clean up truncated text
            loc = re.sub(r'\s+', ' ', loc)
            if len(loc) > 200:
                loc = loc[:200] + "..."
            post.location = loc


# ────────────────────────────────────────────────────────────
# Export: CSV (semikolon-separert, ns3420reader-format)
# ────────────────────────────────────────────────────────────

def export_csv(doc: Document, output_path: Path):
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['Postnr', 'NS3420', 'Emne', 'Beskrivelse',
                         'Mengde', 'Enhet', 'Pris', 'Sum'])
        for p in doc.posts:
            writer.writerow([
                p.post_number,
                p.ns3420_code,
                p.subject,
                p.description,
                f"{p.quantity:.2f}" if p.quantity else "",
                p.unit,
                f"{p.unit_price:.2f}" if p.unit_price is not None else "",
                f"{p.total_price:.2f}" if p.total_price is not None else "",
            ])
    print(f"  CSV: {output_path} ({len(doc.posts)} poster)")


# ────────────────────────────────────────────────────────────
# Export: Excel (.xlsx) med flere ark
# ────────────────────────────────────────────────────────────

def export_excel(doc: Document, output_path: Path):
    if not HAS_OPENPYXL:
        print("  FEIL: openpyxl ikke installert (pip install openpyxl)")
        return

    wb = openpyxl.Workbook()

    # Styles
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_fill2 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Border(
        left=XlSide(style='thin'), right=XlSide(style='thin'),
        top=XlSide(style='thin'), bottom=XlSide(style='thin'),
    )
    num_fmt = '#,##0.00'
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    def styled_header(ws, headers, fill=None):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = fill or hdr_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin

    # ── Ark 1: Postliste (ns3420reader-kompatibel) ──
    ws1 = wb.active
    ws1.title = "Postliste"
    headers1 = ['Postnr', 'NS3420', 'Emne', 'Beskrivelse', 'Mengde', 'Enhet', 'Pris', 'Sum']
    styled_header(ws1, headers1)

    for r, p in enumerate(doc.posts, 2):
        vals = [p.post_number, p.ns3420_code, p.subject, p.description,
                p.quantity if p.quantity else None, p.unit,
                p.unit_price, p.total_price]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = thin
            if c in (5, 7, 8) and v is not None:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal='right')
        # Alternating row color
        if r % 2 == 0:
            for c in range(1, 9):
                ws1.cell(row=r, column=c).fill = alt_fill

    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 24
    ws1.column_dimensions['C'].width = 32
    ws1.column_dimensions['D'].width = 52
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 8
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 14
    ws1.auto_filter.ref = f"A1:H{len(doc.posts) + 1}"
    ws1.freeze_panes = "A2"

    # ── Ark 2: Prismatrise (for tilbudsvurdering) ──
    ws2 = wb.create_sheet("Prismatrise")
    headers2 = ['Postnr', 'Beskrivelse', 'Enhet', 'Mengde',
                'Tilbud 1', 'Tilbud 2', 'Tilbud 3']
    styled_header(ws2, headers2, hdr_fill2)

    for r, p in enumerate(doc.posts, 2):
        ws2.cell(row=r, column=1, value=p.post_number).border = thin
        ws2.cell(row=r, column=2, value=p.description).border = thin
        ws2.cell(row=r, column=3, value=p.unit).border = thin
        qty_cell = ws2.cell(row=r, column=4, value=p.quantity if p.quantity else None)
        qty_cell.border = thin
        if p.quantity:
            qty_cell.number_format = num_fmt
        for c in range(5, 8):
            cell = ws2.cell(row=r, column=c, value=None)
            cell.border = thin
            cell.number_format = num_fmt

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 8
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 14
    ws2.freeze_panes = "A2"

    # ── Ark 3: Kapitteloversikt ──
    ws3 = wb.create_sheet("Kapitteloversikt")
    headers3 = ['Kapittel', 'Tittel', 'Antall poster', 'Kapittelsum']
    styled_header(ws3, headers3)

    ch_counts: Dict[str, int] = {}
    for p in doc.posts:
        ch_counts[p.chapter_code] = ch_counts.get(p.chapter_code, 0) + 1

    for r, (code, title) in enumerate(sorted(doc.chapters.items()), 2):
        ws3.cell(row=r, column=1, value=code).border = thin
        ws3.cell(row=r, column=2, value=title).border = thin
        ws3.cell(row=r, column=3, value=ch_counts.get(code, 0)).border = thin
        cell = ws3.cell(row=r, column=4, value=None)
        cell.border = thin
        cell.number_format = num_fmt

    # Total row
    total_row = len(doc.chapters) + 2
    ws3.cell(row=total_row, column=1, value="TOTALT").font = Font(bold=True)
    ws3.cell(row=total_row, column=3, value=len(doc.posts)).font = Font(bold=True)

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 16

    # ── Ark 4: Full beskrivelse ──
    ws4 = wb.create_sheet("Full beskrivelse")
    headers4 = ['Postnr', 'NS3420', 'Beskrivelse', 'Enhet', 'Mengde',
                'Lokalisering', 'Kapittel', 'Full tekst']
    styled_header(ws4, headers4)

    for r, p in enumerate(doc.posts, 2):
        ws4.cell(row=r, column=1, value=p.post_number).border = thin
        ws4.cell(row=r, column=2, value=p.ns3420_code).border = thin
        ws4.cell(row=r, column=3, value=p.description).border = thin
        ws4.cell(row=r, column=4, value=p.unit).border = thin
        qty = ws4.cell(row=r, column=5, value=p.quantity if p.quantity else None)
        qty.border = thin
        if p.quantity:
            qty.number_format = num_fmt
        ws4.cell(row=r, column=6, value=p.location).border = thin
        ws4.cell(row=r, column=7, value=f"{p.chapter_code} {p.chapter_title}").border = thin
        ws4.cell(row=r, column=8, value=p.full_text[:5000]).border = thin

    for col, w in zip('ABCDEFGH', [14, 22, 42, 8, 12, 42, 30, 80]):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A2"

    # ── Ark 5: Statistikk ──
    ws5 = wb.create_sheet("Statistikk")
    title_font = Font(bold=True, size=14)
    ws5.cell(row=1, column=1, value="NS 3420 Import - Statistikk").font = title_font
    ws5.cell(row=1, column=1).fill = hdr_fill
    ws5.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws5.merge_cells("A1:B1")

    stats = [
        ("Prosjekt", doc.project_name),
        ("Dokument", doc.document_name),
        ("Dato", doc.date),
        ("Oppdragsgiver", doc.client),
        ("Radgiver", doc.author),
        ("", ""),
        ("Totalt antall poster", len(doc.posts)),
        ("Antall kapitler", len(doc.chapters)),
    ]

    # Unit distribution
    unit_counts: Dict[str, int] = {}
    for p in doc.posts:
        u = p.unit or "(ukjent)"
        unit_counts[u] = unit_counts.get(u, 0) + 1

    stats.append(("", ""))
    stats.append(("Enhetsfordeling", "Antall"))
    for unit, count in sorted(unit_counts.items(), key=lambda x: -x[1]):
        stats.append((f"  {unit}", count))

    # Chapter distribution
    stats.append(("", ""))
    stats.append(("Kapittelfordeling", "Antall"))
    for code in sorted(doc.chapters):
        count = ch_counts.get(code, 0)
        stats.append((f"  {code} {doc.chapters[code]}", count))

    for r, (label, value) in enumerate(stats, 3):
        ws5.cell(row=r, column=1, value=label).font = Font(bold=True) if not label.startswith("  ") else Font()
        ws5.cell(row=r, column=2, value=value)

    ws5.column_dimensions['A'].width = 50
    ws5.column_dimensions['B'].width = 20

    wb.save(output_path)
    print(f"  Excel: {output_path} ({len(doc.posts)} poster, 5 ark)")


# ────────────────────────────────────────────────────────────
# Export: NS 3459 XML
# ────────────────────────────────────────────────────────────

def export_xml(doc: Document, output_path: Path):
    ns = "http://www.standard.no/ns3459"
    root = ET.Element("NS3459", xmlns=ns)

    meta = ET.SubElement(root, "Metadata")
    ET.SubElement(meta, "Prosjekt").text = doc.project_name
    ET.SubElement(meta, "Dokument").text = doc.document_name
    ET.SubElement(meta, "Dato").text = doc.date
    ET.SubElement(meta, "Valuta").text = "NOK"
    if doc.client:
        ET.SubElement(meta, "Oppdragsgiver").text = doc.client
    if doc.author:
        ET.SubElement(meta, "Konsulent").text = doc.author

    for post in doc.posts:
        p = ET.SubElement(root, "Post")
        ET.SubElement(p, "ID").text = str(uuid.uuid4())
        ET.SubElement(p, "Postnr").text = post.post_number

        if post.ns3420_code:
            kode = ET.SubElement(p, "Kode")
            ET.SubElement(kode, "ID").text = post.ns3420_code

        ET.SubElement(p, "Beskrivelse").text = post.description

        if post.subject:
            ET.SubElement(p, "Emne").text = post.subject

        prisinfo = ET.SubElement(p, "Prisinfo")
        ET.SubElement(prisinfo, "Mengde").text = str(post.quantity)
        ET.SubElement(prisinfo, "Enhet").text = post.unit
        if post.unit_price is not None:
            ET.SubElement(prisinfo, "Enhetspris").text = str(post.unit_price)
        if post.total_price is not None:
            ET.SubElement(prisinfo, "Sum").text = str(post.total_price)

    xml_str = ET.tostring(root, encoding='unicode')
    try:
        dom = minidom.parseString(xml_str)
        pretty = dom.toprettyxml(indent="  ", encoding="UTF-8")
    except Exception:
        pretty = f'<?xml version="1.0" encoding="UTF-8"?>\n{xml_str}'.encode('utf-8')

    with open(output_path, 'wb') as f:
        f.write(pretty)
    print(f"  XML: {output_path} ({len(doc.posts)} poster)")


# ────────────────────────────────────────────────────────────
# Export: JSON (for integrasjon med andre verktoy)
# ────────────────────────────────────────────────────────────

def export_json(doc: Document, output_path: Path):
    import json

    data = {
        "metadata": {
            "project_name": doc.project_name,
            "document_name": doc.document_name,
            "date": doc.date,
            "client": doc.client,
            "author": doc.author,
            "total_posts": len(doc.posts),
            "chapters": doc.chapters,
        },
        "posts": [
            {
                "postnr": p.post_number,
                "ns3420": p.ns3420_code,
                "emne": p.subject,
                "beskrivelse": p.description,
                "mengde": p.quantity,
                "enhet": p.unit,
                "pris": p.unit_price,
                "sum": p.total_price,
                "lokalisering": p.location,
                "kapittel": p.chapter_code,
                "kapittel_tittel": p.chapter_title,
                "side": p.page,
            }
            for p in doc.posts
        ],
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  JSON: {output_path} ({len(doc.posts)} poster)")


# ────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="NS 3420 PDF Import -> ns3420reader-kompatibelt regneark",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Eksempler:
  python ns3420pdfimport.py beskrivelse.pdf
  python ns3420pdfimport.py beskrivelse.pdf --format xlsx
  python ns3420pdfimport.py beskrivelse.pdf --format csv --output resultat
  python ns3420pdfimport.py extracted.txt --format all
  python ns3420pdfimport.py beskrivelse.pdf --format json
        """
    )
    parser.add_argument('input', help='PDF-fil eller allerede ekstrahert .txt-fil')
    parser.add_argument('--format', '-f', choices=['xlsx', 'csv', 'xml', 'json', 'all'],
                        default='all', help='Eksportformat (standard: all)')
    parser.add_argument('--output', '-o', help='Utdatafilnavn (uten endelse)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Vis detaljert output')

    args = parser.parse_args()
    input_path = Path(args.input).resolve()

    if not input_path.exists():
        print(f"FEIL: Filen finnes ikke: {input_path}")
        sys.exit(1)

    output_base = args.output or input_path.stem
    output_dir = Path.cwd()

    # ── Parse ──
    print(f"\n[1/2] Parser: {input_path.name}")

    if input_path.suffix == '.pdf' and HAS_FITZ:
        # Coordinate-based parsing (best accuracy)
        coord_parser = CoordinateParser()
        doc = coord_parser.parse(input_path)
    elif input_path.suffix == '.txt':
        # Text-based parsing (fallback for pre-extracted text)
        text = input_path.read_text(encoding='utf-8')
        print(f"  {len(text):,} tegn (tekst-modus)")
        ns_parser = NS3420Parser()
        doc = ns_parser.parse(text)
    else:
        # PDF without PyMuPDF -> use text extraction fallback
        print("  PyMuPDF ikke tilgjengelig, bruker tekst-ekstraksjon...")
        text = extract_text(input_path)
        ns_parser = NS3420Parser()
        doc = ns_parser.parse(text)

    real_posts = [p for p in doc.posts if p.post_number]
    print(f"  Funnet: {len(real_posts)} poster i {len(doc.chapters)} kapitler")
    if doc.project_name:
        print(f"  Prosjekt: {doc.project_name}")

    if not doc.posts:
        print("\nADVARSEL: Ingen poster funnet. Sjekk at filen er en NS 3420 beskrivelse.")
        sys.exit(1)

    if args.verbose:
        print("\n  Kapitler:")
        for code in sorted(doc.chapters):
            count = sum(1 for p in real_posts if p.chapter_code == code)
            print(f"    {code} {doc.chapters[code]} ({count} poster)")

        print(f"\n  Forste 15 poster:")
        for p in real_posts[:15]:
            desc_short = p.description.split('\n')[0][:45]
            print(f"    {p.post_number:16s} {p.ns3420_code:22s} "
                  f"{p.unit:5s} {p.quantity:>10.2f}  {desc_short}")

        unit_counts: Dict[str, int] = {}
        for p in real_posts:
            u = p.unit or "(ukjent)"
            unit_counts[u] = unit_counts.get(u, 0) + 1
        print(f"\n  Enhetsfordeling:")
        for unit, count in sorted(unit_counts.items(), key=lambda x: -x[1]):
            print(f"    {unit:8s} {count:4d} poster")

    # ── Export ──
    print(f"\n[2/2] Eksporterer...")
    formats = ['xlsx', 'csv', 'xml', 'json'] if args.format == 'all' else [args.format]

    for fmt in formats:
        out_path = output_dir / f"{output_base}.{fmt}"
        if fmt == 'xlsx':
            export_excel(doc, out_path)
        elif fmt == 'csv':
            export_csv(doc, out_path)
        elif fmt == 'xml':
            export_xml(doc, out_path)
        elif fmt == 'json':
            export_json(doc, out_path)

    print(f"\nFerdig! {len(doc.posts)} poster eksportert.")


if __name__ == '__main__':
    main()
